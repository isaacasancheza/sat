#!/usr/bin/env -S uv run
import ast
from functools import cache
from pathlib import Path
from typing import Generator, NotRequired, TypedDict

import astor
from openpyxl import load_workbook

workbook = Path('../assets/catCFDI_V_4_20250313.xlsx')
package_root = Path('../src/sat')
catalogos_module = package_root / 'catalogos'


class Catalogo(TypedDict):
    module_name: str
    enum_name: str
    enum_values_prefix: NotRequired[str]
    use_enum_value_as_name: NotRequired[bool]
    enum_values: Generator[tuple[str, str]]


def create_enum(
    *,
    enum_name: str,
    enum_values: Generator[tuple[str, str]],
    enum_values_prefix: str | None = None,
    use_enum_value_as_name: bool = False,
):
    body = []
    for key, value in enum_values:
        name = key
        if use_enum_value_as_name:
            name = value.upper()
        if enum_values_prefix is not None:
            name = f'{enum_values_prefix}{name}'
        body.append(
            ast.Assign(
                targets=[
                    ast.Name(
                        id=name,
                        ctx=ast.Store(),
                    ),
                ],
                value=ast.Constant(
                    value=key,
                ),
            )
        )
        body.append(
            ast.Expr(
                value=ast.Constant(
                    value=value,
                ),
            ),
        )
    enum = ast.ClassDef(
        name=enum_name,
        bases=[
            ast.Name(
                id='StrEnum',
                ctx=ast.Load(),
            ),
        ],
        body=body,
    )

    # import enum.StreEnum
    import_enum = ast.ImportFrom(
        module='enum',
        names=[
            ast.alias(
                name='StrEnum',
            ),
        ],
        level=0,
    )

    # create module
    module = ast.Module(
        body=[
            import_enum,
            enum,
        ],
    )

    # convert ast to source code
    source_code = astor.to_source(module)

    return source_code


@cache
def get_workbook():
    return load_workbook(workbook)


def int02d(
    value: str,
    /,
):
    return f'{int(value):02d}'


def int03d(
    value: str,
    /,
):
    return f'{int(value):03d}'


def get_data(
    *,
    min_row: int,
    worksheet: str,
):
    wb = get_workbook()
    for row in wb[worksheet].iter_rows(
        min_col=1,
        max_col=2,
        min_row=min_row,
    ):
        key, value = row
        if key.value is None:
            continue
        key = str(key.value).strip()
        value = str(value.value).strip()
        yield key, value


def get_forma_pago():
    return (
        (int02d(key), value)
        for key, value in get_data(
            min_row=7,
            worksheet='c_FormaPago',
        )
    )


def get_moneda():
    return (
        (key, value)
        for key, value in get_data(
            min_row=6,
            worksheet='c_Moneda',
        )
    )


def get_tipo_de_comprobante():
    return (
        (key, value)
        for key, value in get_data(
            min_row=6,
            worksheet='c_TipoDeComprobante',
        )
    )


def get_exportacion():
    return (
        (int02d(key), value)
        for key, value in get_data(
            min_row=6,
            worksheet='c_Exportacion',
        )
    )


def get_metodo_pago():
    return (
        (key, value)
        for key, value in get_data(
            min_row=7,
            worksheet='c_MetodoPago',
        )
    )


def get_peridiocidad():
    return (
        (int02d(key), value)
        for key, value in get_data(
            min_row=7,
            worksheet='c_Periodicidad',
        )
    )


def get_meses():
    return (
        (int02d(key), value)
        for key, value in get_data(
            min_row=6,
            worksheet='c_Meses',
        )
    )


def get_tipo_relacion():
    return (
        (int02d(key), value)
        for key, value in get_data(
            min_row=6,
            worksheet='c_TipoRelacion',
        )
    )


def get_regimen_fiscal():
    return (
        (key, value)
        for key, value in get_data(
            min_row=7,
            worksheet='c_RegimenFiscal',
        )
    )


def get_pais():
    return (
        (key, value)
        for key, value in get_data(
            min_row=6,
            worksheet='c_Pais',
        )
    )


def get_uso_cfdi():
    return (
        (key, value)
        for key, value in get_data(
            min_row=7,
            worksheet='c_UsoCFDI',
        )
    )


def get_clave_prod_serv():
    return (
        (key, value)
        for key, value in get_data(
            min_row=6,
            worksheet='c_ClaveProdServ',
        )
    )


def get_clave_unidad():
    return (
        (key, value)
        for key, value in get_data(
            min_row=7,
            worksheet='c_ClaveUnidad',
        )
    )


def get_objeto_imp():
    return (
        (int02d(key), value)
        for key, value in get_data(
            min_row=6,
            worksheet='c_ObjetoImp',
        )
    )


def get_impuesto():
    return (
        (int03d(key), value)
        for key, value in get_data(
            min_row=6,
            worksheet='c_Impuesto',
        )
    )


def get_aduana():
    return (
        (int02d(key), value)
        for key, value in get_data(
            min_row=6,
            worksheet='c_Aduana',
        )
    )


def get_tipo_factor():
    wb = get_workbook()
    for row in wb['c_TipoFactor'].iter_rows(
        min_row=6,
        min_col=1,
        max_col=1,
    ):
        (value,) = row
        if value.value is None:
            continue
        key = str(value.value).strip()
        value = str(value.value).strip()
        yield key, value


catalogos: list[Catalogo] = [
    {
        'module_name': 'forma_pago',
        'enum_name': 'FormaPago',
        'enum_values_prefix': 'FP_',
        'enum_values': get_forma_pago(),
    },
    {
        'module_name': 'moneda',
        'enum_name': 'Moneda',
        'enum_values': get_moneda(),
        'enum_values_prefix': 'M_',
    },
    {
        'module_name': 'tipo_de_comprobante',
        'enum_name': 'TipoDeComprobante',
        'enum_values': get_tipo_de_comprobante(),
        'use_enum_value_as_name': True,
        'enum_values_prefix': 'TC_',
    },
    {
        'module_name': 'exportacion',
        'enum_name': 'Exporacion',
        'enum_values_prefix': 'E_',
        'enum_values': get_exportacion(),
    },
    {
        'module_name': 'metodo_pago',
        'enum_name': 'MetodoPago',
        'enum_values': get_metodo_pago(),
        'enum_values_prefix': 'MP_',
    },
    {
        'module_name': 'peridiocidad',
        'enum_name': 'Peridiocidad',
        'enum_values': get_peridiocidad(),
        'use_enum_value_as_name': True,
        'enum_values_prefix': 'P_',
    },
    {
        'module_name': 'meses',
        'enum_name': 'Meses',
        'enum_values_prefix': 'M_',
        'enum_values': get_meses(),
    },
    {
        'module_name': 'tipo_relacion',
        'enum_name': 'TipoRelacion',
        'enum_values_prefix': 'TR_',
        'enum_values': get_tipo_relacion(),
    },
    {
        'module_name': 'regimen_fiscal',
        'enum_name': 'RegimenFiscal',
        'enum_values_prefix': 'RF_',
        'enum_values': get_regimen_fiscal(),
    },
    {
        'module_name': 'pais',
        'enum_name': 'Pais',
        'enum_values_prefix': 'P_',
        'enum_values': get_pais(),
    },
    {
        'module_name': 'uso_cfdi',
        'enum_name': 'UsoCFDI',
        'enum_values_prefix': 'UC_',
        'enum_values': get_uso_cfdi(),
    },
    # too big
    # {
    #     'module_name': 'clave_prod_serv',
    #     'enum_name': 'ClaveProdServ',
    #     'enum_values_prefix': 'CPS_',
    #     'enum_values': get_clave_prod_serv(),
    # },
    {
        'module_name': 'clave_unidad',
        'enum_name': 'ClaveUnidad',
        'enum_values_prefix': 'CU_',
        'enum_values': get_clave_unidad(),
    },
    {
        'module_name': 'objeto_imp',
        'enum_name': 'ObjetoImp',
        'enum_values_prefix': 'OI_',
        'enum_values': get_objeto_imp(),
    },
    {
        'module_name': 'impuesto',
        'enum_name': 'Impuesto',
        'enum_values_prefix': 'I_',
        'enum_values': get_impuesto(),
    },
    {
        'module_name': 'aduana',
        'enum_name': 'Aduana',
        'enum_values_prefix': 'A_',
        'enum_values': get_aduana(),
    },
    {
        'module_name': 'tipo_factor',
        'enum_name': 'TipoFactor',
        'enum_values': get_tipo_factor(),
        'enum_values_prefix': 'TF_',
        'use_enum_value_as_name': True,
    },
]

package_root.mkdir(exist_ok=True)

package_init = package_root / '__init__.py'
package_init.touch(exist_ok=True)

catalogos_module.mkdir(exist_ok=True)

imports: list[ast.stmt] = []
for catalogo in catalogos:
    enum_name = catalogo['enum_name']
    module_name = catalogo['module_name']
    source_code = create_enum(
        enum_name=enum_name,
        enum_values=catalogo['enum_values'],
        enum_values_prefix=catalogo.get('enum_values_prefix'),
        use_enum_value_as_name=catalogo.get('use_enum_value_as_name', False),
    )

    imprt = ast.ImportFrom(
        module=module_name,
        names=[
            ast.alias(
                name=enum_name,
            ),
        ],
        level=1,
    )

    imports.append(imprt)

    module_path = catalogos_module / f'{module_name}.py'
    with open(module_path, 'w') as f:
        f.write(source_code)


# import enums
init_module = ast.Module(body=imports)
init_source_code = astor.to_source(init_module)
with open(catalogos_module / '__init__.py', 'w', encoding='utf-8') as f:
    f.write(init_source_code)
